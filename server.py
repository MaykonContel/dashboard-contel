import os
import io
import json
import traceback
from datetime import datetime, date
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_JSON = os.path.join(BASE_DIR, "data_store.json")
INDEX_HTML = os.path.join(BASE_DIR, "dashboard.html")
UPLOAD_XLSX = os.path.join(BASE_DIR, "_base_atual.xlsx")
PORT = int(os.environ.get("PORT", 8000))


def to_iso(v):
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time()).isoformat()
    return v


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    sheet_plan = wb["PLANEJAMENTO"]
    sheet_cron = wb["CRONOGRAMA"]
    sheet_equipe = wb["EQUIPE"]
    sheet_etapa = wb["ETAPA OBRA"]
    sheet_conf = wb["CONFIGURAÇÕES"] if "CONFIGURAÇÕES" in wb.sheetnames else None

    # PLANEJAMENTO
    headers = [cell.value if cell.value is not None else "" for cell in next(sheet_plan.iter_rows(min_row=1, max_row=1))]
    parsed = []
    plan_ids = []
    for row in sheet_plan.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        item = {}
        for h, v in zip(headers, row):
            item[str(h)] = to_iso(v)
        parsed.append(item)
        site = str(row[0]).strip().upper() if row and row[0] not in (None, "") else ""
        plan_ids.append(site)

    # CRONOGRAMA
    cron_map = {}
    inicio_map = {}
    status_prazo_map = {}
    hoje = datetime.now().date()

    for row in sheet_cron.iter_rows(min_row=2, values_only=True):
        site = row[0]
        if not site:
            continue
        key = str(site).strip().upper()
        dt_i = row[5] if len(row) > 5 else None
        dt_f = row[6] if len(row) > 6 else None
        checkbox = str(row[9] if len(row) > 9 and row[9] is not None else "").strip().upper()

        if isinstance(dt_i, (datetime, date)):
            inicio_map[key] = to_iso(dt_i)
        if isinstance(dt_f, (datetime, date)):
            cron_map[key] = to_iso(dt_f)

        data_fim = dt_f.date() if isinstance(dt_f, datetime) else dt_f if isinstance(dt_f, date) else None
        if checkbox == "FINALIZADO":
            status_prazo = "✅ CONCLUÍDO"
        elif checkbox == "PARALISADO":
            status_prazo = "⚠️ PARALISADO"
        elif checkbox == "CANCELADO":
            status_prazo = "🚫 CANCELADO"
        elif data_fim and data_fim < hoje:
            status_prazo = "⏰ ATRASADO"
        else:
            status_prazo = "🕓 EM ANDAMENTO"
        status_prazo_map[key] = status_prazo

    # CONFIGURAÇÕES / fallback ETAPA OBRA headers
    etapas_list = []
    if sheet_conf is not None:
        started = False
        for row in sheet_conf.iter_rows(values_only=True):
            val = str(row[0] if row and row[0] is not None else "").strip()
            if not val:
                continue
            if "LISTA DE ETAPA" in val.upper():
                started = True
                continue
            if started:
                etapas_list.append(val)

    etapa_headers = [cell.value if cell.value is not None else "" for cell in next(sheet_etapa.iter_rows(min_row=1, max_row=1))]
    if not etapas_list:
        etapas_list = [
            str(h).strip() for h in etapa_headers[1:]
            if str(h).strip() and str(h).strip().lower() not in {"paralisada", "cancelada", "finalizado"}
        ]

    # EQUIPE
    equipe_map = {}
    equipe_rows = list(sheet_equipe.iter_rows(min_row=2, values_only=True))
    for idx, row in enumerate(equipe_rows):
        site = plan_ids[idx] if idx < len(plan_ids) else ""
        if not site:
            continue
        membros = []
        vals = list(row)
        for j in range(3):
            nome = str(vals[1 + j * 2] if len(vals) > 1 + j * 2 and vals[1 + j * 2] is not None else "").strip()
            valor = vals[2 + j * 2] if len(vals) > 2 + j * 2 else None
            try:
                valor = float(valor) if valor not in (None, "") else None
            except Exception:
                valor = None
            if nome:
                membros.append({"nome": nome, "valor": valor})
        if membros:
            equipe_map[site] = membros

    # ETAPA OBRA
    etapa_obra_map = {}
    etapa_rows = list(sheet_etapa.iter_rows(min_row=2, values_only=True))
    for idx, row in enumerate(etapa_rows):
        site = plan_ids[idx] if idx < len(plan_ids) else ""
        if not site:
            continue
        vals = list(row)[1:]
        etapas = []
        for name, val in zip(etapa_headers[1:], vals):
            if str(val if val is not None else "").strip().upper() == "SIM" and str(name).strip():
                etapas.append(str(name).strip())
        if etapas:
            etapa_obra_map[site] = etapas

    payload = {
        "ok": True,
        "parsed": parsed,
        "cronMap": cron_map,
        "inicioMap": inicio_map,
        "etapasList": etapas_list,
        "equipeMap": equipe_map,
        "etapaObraMap": etapa_obra_map,
        "statusPrazoMap": status_prazo_map,
        "updated_at": datetime.now().isoformat()
    }
    return payload


def load_saved_payload():
    if not os.path.exists(DATA_JSON):
        return {}
    with open(DATA_JSON, "r", encoding="utf-8") as f:
        return json.load(f)


def save_payload(payload):
    with open(DATA_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)


class Handler(BaseHTTPRequestHandler):
    server_version = "ContelServer/1.0"

    def _send_json(self, data, status=200):
        raw = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(raw)

    def _send_html(self, content, status=200):
        raw = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(raw)

    def _send_text(self, content, status=200):
        raw = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(raw)

    def log_message(self, fmt, *args):
        print("[%s] %s" % (self.address_string(), fmt % args))

    def do_HEAD(self):
        parsed = urlparse(self.path)
        if parsed.path in ("/", "/dashboard.html"):
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
        elif parsed.path.startswith("/api/"):
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
        else:
            self.send_response(404)
            self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path in ("/", "/dashboard.html"):
            with open(INDEX_HTML, "r", encoding="utf-8") as f:
                self._send_html(f.read())
            return

        if path == "/api/health":
            self._send_json({"ok": True, "service": "contel-dashboard"})
            return

        if path == "/api/data":
            payload = load_saved_payload()
            self._send_json(payload or {"ok": True, "parsed": []})
            return

        self._send_text("Not found", status=404)


    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path != "/api/upload-xlsx":
            self._send_json({"ok": False, "error": "Rota não encontrada."}, status=404)
            return

        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length <= 0:
                self._send_json({"ok": False, "error": "Envio vazio."}, status=400)
                return

            content_type = self.headers.get("Content-Type", "")
            if "multipart/form-data" not in content_type or "boundary=" not in content_type:
                self._send_json({"ok": False, "error": "Formato de envio inválido."}, status=400)
                return

            boundary = content_type.split("boundary=", 1)[1].strip().encode("utf-8")
            body = self.rfile.read(length)

            parts = body.split(b"--" + boundary)
            file_bytes = None
            for part in parts:
                if b'Content-Disposition' not in part or b'name="file"' not in part:
                    continue
                header_end = part.find(b"\r\n\r\n")
                if header_end == -1:
                    continue
                data = part[header_end + 4:]
                if data.endswith(b"\r\n"):
                    data = data[:-2]
                if data.endswith(b"--"):
                    data = data[:-2]
                file_bytes = data
                break

            if not file_bytes:
                self._send_json({"ok": False, "error": "Arquivo não enviado."}, status=400)
                return

            with open(UPLOAD_XLSX, "wb") as f:
                f.write(file_bytes)

            payload = parse_workbook(UPLOAD_XLSX)
            save_payload(payload)
            self._send_json(payload, status=200)

        except Exception as e:
            traceback.print_exc()
            self._send_json({"ok": False, "error": str(e)}, status=500)


if __name__ == "__main__":
    # Garantia de base inicial
    if not os.path.exists(DATA_JSON) and os.path.exists(UPLOAD_XLSX):
        try:
            save_payload(parse_workbook(UPLOAD_XLSX))
        except Exception:
            traceback.print_exc()

    server = ThreadingHTTPServer(("0.0.0.0", PORT), Handler)
    print(f"Servidor rodando na porta {PORT}")
    server.serve_forever()
